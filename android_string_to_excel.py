"""
将 Android string.xml 文件中的文本转换成 Excel 表格并保存到文件。
使用方法：python android_string_to_excel.py -f strings.xml [-o outfile_name]
-f: 要处理的 Android string.xml 文件路径.
-o: 输出文件路径，如果未指定，则默认为 'strings.xlsx'.
"""

import os
import argparse

import openpyxl
from lxml import etree


def printUsage():
    """print usage and exit"""
    usage = '''Usage:
        python android_string_to_excel.py -f strings.xml [-o outfile_name]
        '''
    print(usage)
    exit(-1)


parser = argparse.ArgumentParser(description='Convert Android string.xml to Excel')
parser.add_argument('-f', '--file',
                    type=str,
                    required=True,
                    help='Android string.xml file path')
parser.add_argument('-o', '--output',
                    type=str,
                    help='Output file name, default is "strings.xlsx"')
args = parser.parse_args()

xml_path = args.file
out_file = args.output or 'strings.xlsx'

if not os.path.isfile(xml_path):
    print(f'Error: {xml_path} is not a valid file!')
    printUsage()

try:
    tree = etree.parse(xml_path)
except etree.ParseError as e:
    print(f'Error: cannot parse XML at {xml_path}, {e}')
    exit(-1)

root = tree.getroot()
workbook = openpyxl.Workbook()
sheet = workbook.active
col_idx = 1
headers = ['key', 'value']
for col_idx, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_idx, value=header)

for row_idx, child in enumerate(root, 2):
    key = child.attrib['name']
    value = child.text.strip() if child.text else ''
    sheet.cell(row=row_idx, column=1, value=key)
    sheet.cell(row=row_idx, column=2, value=value)

workbook.save(out_file)
print(f'{xml_path} converted and saved to {out_file}.')
